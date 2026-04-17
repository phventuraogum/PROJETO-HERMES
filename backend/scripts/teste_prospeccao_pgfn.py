"""
TESTE COMPLETO DO SISTEMA -- 100 empresas SP com divida PGFN <= R$10M
Pipeline: DB query -> OpenCNPJ -> scraping web (ddgs) -> socios -> Excel

Uso:
    python scripts/teste_prospeccao_pgfn.py
"""
from __future__ import annotations

import asyncio
import json
import os
import re
import sys
import time
from datetime import datetime
from pathlib import Path

# Carrega .env do projeto (raiz pai do backend)
_ROOT = Path(__file__).resolve().parents[2]
_BACKEND = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(_BACKEND))

for _env in [_ROOT / ".env", _BACKEND / ".env", _BACKEND / ".env.local"]:
    if _env.exists():
        from dotenv import load_dotenv
        load_dotenv(_env, override=True)

# Forca DuckDuckGo como provider de busca (sem API key)
os.environ.setdefault("SEARCH_PROVIDER_ORDER", "ddgs,brave,bing")

import duckdb
import httpx

CNPJ_DB   = os.getenv("HERMES_DUCKDB_PATH", "G:/icp_radar/dados_receita/cnpj.duckdb")
SAIDA_XLS = str(_BACKEND / "scripts" / f"hermes_pgfn_sp_{datetime.now():%Y%m%d_%H%M}.xlsx")
LIMITE    = 100
CONCORRENCIA = 15   # OpenCNPJ suporta ate 50 req/s


# -----------------------------------------------------------------------------
# 1. SELECAO DE EMPRESAS (mix de setores)
# -----------------------------------------------------------------------------

def selecionar_100_empresas(con: duckdb.DuckDBPyConnection):
    """
    Busca 100 empresas variadas: ~25 por faixa de porte/divida para diversidade.
    Usa mix de setores para teste mais representativo.
    """
    print("Selecionando 100 empresas SP com divida PGFN...")

    # Faixas para garantir diversidade
    faixas = [
        (5_000_000, 10_000_000, 30),   # grandes dividas
        (1_000_000,  5_000_000, 30),   # medias
        (100_000,    1_000_000, 25),   # menores
        (1,          100_000,   15),   # micro
    ]

    todos = []
    seen = set()

    for dmin, dmax, lim in faixas:
        rows = con.execute("""
            SELECT
                b.cnpj,
                b.RAZAO_SOCIAL          AS razao_social,
                b.NOME_FANTASIA         AS nome_fantasia,
                b.CNAE_PRINCIPAL        AS cnae_principal,
                b.cnae_descricao,
                b.PORTE_EMPRESA         AS porte_cod,
                b.CAPITAL_SOCIAL_NUM    AS capital_social,
                b.email_receita,
                b.telefone_receita,
                b.cidade_nome           AS cidade,
                b.UF                    AS uf,
                b.LOGRADOURO            AS logradouro,
                b.NUMERO                AS numero,
                b.CEP                   AS cep,
                b.DATA_INICIO_ATIVIDADE AS data_abertura,
                p.divida_total          AS divida_pgfn,
                p.qtd_inscricoes        AS inscricoes_pgfn
            FROM vw_prospeccao_base AS b
            INNER JOIN pgfn_dividas AS p ON p.cnpj = b.cnpj
            WHERE b.UF = 'SP'
              AND p.divida_total BETWEEN ? AND ?
              AND p.divida_total > 0
            ORDER BY b.CAPITAL_SOCIAL_NUM DESC NULLS LAST
            LIMIT ?
        """, [dmin, dmax, lim * 3]).fetchdf()

        for _, r in rows.iterrows():
            if r["cnpj"] not in seen and len(todos) < LIMITE:
                seen.add(r["cnpj"])
                todos.append(r.to_dict())
            if len(todos) >= LIMITE:
                break

    print(f"  Selecionadas: {len(todos)} empresas")
    return todos


# -----------------------------------------------------------------------------
# 2. ENRIQUECIMENTO OpenCNPJ (free, sem key)
# -----------------------------------------------------------------------------

async def enriquecer_opencnpj_lote(empresas: list, sem: asyncio.Semaphore) -> dict:
    from enrichment_opencnpj import consultar_opencnpj

    async def _um(emp):
        async with sem:
            cnpj = str(emp["cnpj"])
            try:
                return cnpj, await consultar_opencnpj(cnpj)
            except Exception:
                return cnpj, {}

    tasks = [_um(e) for e in empresas]
    resultados = await asyncio.gather(*tasks, return_exceptions=False)
    return {cnpj: dados for cnpj, dados in resultados}


# -----------------------------------------------------------------------------
# 3. BUSCA WEB + SCRAPING (DuckDuckGo -> site -> contatos)
# -----------------------------------------------------------------------------

async def buscar_site_empresa(
    razao: str,
    cidade: str,
    cnae_desc: str,
    sem: asyncio.Semaphore,
    timeout: float = 6.0,
) -> dict:
    """
    DuckDuckGo search para encontrar site e contatos. Nao requer API key.
    """
    async with sem:
        nome_busca = razao.split()[0] if razao else ""
        query = f'"{razao}" site empresa {cidade} SP contato'
        try:
            from duckduckgo_search import DDGS
            with DDGS(timeout=timeout) as ddgs:
                results = list(ddgs.text(query, max_results=3))
            if results:
                href = results[0].get("href", "")
                body = results[0].get("body", "")
                # Extrai email do snippet
                emails = re.findall(r"[\w.+-]+@[\w-]+\.[a-z]{2,}", body.lower())
                email = emails[0] if emails else None
                return {"site": href or None, "email_web": email, "snippet": body[:200]}
        except Exception:
            pass
        return {}


async def scraping_lote(empresas: list) -> dict:
    sem = asyncio.Semaphore(8)
    tasks = []
    for emp in empresas:
        tasks.append(buscar_site_empresa(
            razao=str(emp.get("razao_social") or ""),
            cidade=str(emp.get("cidade") or "SP"),
            cnae_desc=str(emp.get("cnae_descricao") or ""),
            sem=sem,
        ))
    resultados = await asyncio.gather(*tasks, return_exceptions=True)
    out = {}
    for emp, res in zip(empresas, resultados):
        if isinstance(res, dict):
            out[str(emp["cnpj"])] = res
    return out


# -----------------------------------------------------------------------------
# 4. SOCIOS do DuckDB (tabela socios da Receita)
# -----------------------------------------------------------------------------

def buscar_socios_db(con: duckdb.DuckDBPyConnection, cnpjs: list) -> dict:
    if not cnpjs:
        return {}
    phs = ", ".join("?" * len(cnpjs))
    try:
        df = con.execute(f"""
            SELECT
                LPAD(s.CNPJ_BASICO, 8, '0') AS cnpj_basico,
                s.NOME_SOCIO                 AS nome,
                s.QUALIFICACAO_SOCIO         AS qualificacao,
                s.DATA_ENTRADA_SOCIEDADE     AS data_entrada
            FROM socios s
            WHERE LPAD(s.CNPJ_BASICO, 8, '0') IN ({phs})
            LIMIT 500
        """, [c[:8] for c in cnpjs]).fetchdf()

        socios_map: dict = {}
        for _, r in df.iterrows():
            key = r["cnpj_basico"]
            socios_map.setdefault(key, []).append({
                "nome": (r["nome"] or "").strip().title(),
                "qualificacao": r["qualificacao"],
                "data_entrada": r["data_entrada"],
            })

        # Mapeia cnpj_basico -> cnpj_completo
        result = {}
        for cnpj in cnpjs:
            basico = cnpj[:8]
            result[cnpj] = socios_map.get(basico, [])
        return result
    except Exception as exc:
        print(f"  Aviso socios: {exc}")
        return {}


# -----------------------------------------------------------------------------
# 5. MESCLA DADOS
# -----------------------------------------------------------------------------

PORTE_LABEL = {"01": "ME", "03": "EPP", "05": "Medio/Grande"}

def mesclar(emp: dict, opencnpj: dict, web: dict, socios: list) -> dict:
    cnpj = str(emp["cnpj"])

    # Contatos: prioridade web > opencnpj > receita
    email = (
        (opencnpj.get("email_receita") or "").strip().lower() or
        web.get("email_web") or
        str(emp.get("email_receita") or "")
    ) or None

    tel_oc = opencnpj.get("telefone") or None
    tel_rec = str(emp.get("telefone_receita") or "").strip() or None
    telefone = tel_oc or tel_rec

    site = web.get("site") or None

    # Socios (max 3)
    socios_str = " | ".join(
        f"{s['nome']} ({s['qualificacao']})" for s in socios[:3]
    ) if socios else ""

    # Sócios do OpenCNPJ (QSA)
    qsa = opencnpj.get("socios_qsa") or []
    socios_qsa_str = " | ".join(
        f"{s['nome']} ({s.get('qualificacao','')})"
        for s in qsa[:3]
    ) if qsa else socios_str

    porte = PORTE_LABEL.get(str(emp.get("porte_cod") or "").strip(), str(emp.get("porte_cod") or ""))

    return {
        "CNPJ":             cnpj,
        "Razao Social":     emp.get("razao_social") or "",
        "Nome Fantasia":    emp.get("nome_fantasia") or opencnpj.get("nome_fantasia") or "",
        "CNAE":             emp.get("cnae_principal") or "",
        "Setor":            emp.get("cnae_descricao") or opencnpj.get("cnae_descricao") or "",
        "Porte":            porte,
        "Capital Social":   emp.get("capital_social") or 0,
        "Divida PGFN (R$)": round(float(emp.get("divida_pgfn") or 0), 2),
        "Inscricoes PGFN":  int(emp.get("inscricoes_pgfn") or 0),
        "Email":            email or "",
        "Telefone":         telefone or "",
        "Site":             site or "",
        "Cidade":           emp.get("cidade") or "",
        "UF":               emp.get("uf") or "",
        "Logradouro":       f"{emp.get('logradouro') or ''}, {emp.get('numero') or ''}".strip(", "),
        "CEP":              emp.get("cep") or "",
        "Abertura":         emp.get("data_abertura") or "",
        "Socios":           socios_qsa_str,
        "Fonte Contato":    (
            "OpenCNPJ+Web" if (email and site) else
            "OpenCNPJ" if email else
            "Receita" if tel_rec else
            "--"
        ),
    }


# -----------------------------------------------------------------------------
# 6. PERSISTENCIA EM empresas_enriquecidas
# -----------------------------------------------------------------------------

def persistir(con_rw: duckdb.DuckDBPyConnection, empresas_mescladas: list[dict]):
    salvos = 0
    for emp in empresas_mescladas:
        cnpj = emp["CNPJ"]
        site = emp["Site"] or None
        email = emp["Email"] or None
        telefone = emp["Telefone"] or None
        socios_json = emp["Socios"] or None

        if not any([site, email, telefone]):
            continue

        try:
            con_rw.execute("""
                INSERT INTO empresas_enriquecidas (
                    cnpj, site, email_enriquecido,
                    telefone_enriquecido, outras_informacoes
                ) VALUES (?, ?, ?, ?, ?)
                ON CONFLICT (cnpj) DO UPDATE SET
                    site                 = COALESCE(excluded.site, site),
                    email_enriquecido    = COALESCE(excluded.email_enriquecido, email_enriquecido),
                    telefone_enriquecido = COALESCE(excluded.telefone_enriquecido, telefone_enriquecido),
                    outras_informacoes   = COALESCE(excluded.outras_informacoes, outras_informacoes)
            """, [cnpj, site, email, telefone, socios_json])
            salvos += 1
        except Exception as exc:
            print(f"  Erro ao persistir {cnpj}: {exc}")

    return salvos


# -----------------------------------------------------------------------------
# 7. EXPORT EXCEL (openpyxl)
# -----------------------------------------------------------------------------

def exportar_excel(dados: list[dict], caminho: str):
    try:
        import openpyxl
        from openpyxl.styles import (Font, PatternFill, Alignment,
                                      Border, Side, numbers)
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("  openpyxl nao instalado -- exportando CSV")
        import csv
        csv_path = caminho.replace(".xlsx", ".csv")
        with open(csv_path, "w", newline="", encoding="utf-8-sig") as f:
            if dados:
                writer = csv.DictWriter(f, fieldnames=dados[0].keys())
                writer.writeheader()
                writer.writerows(dados)
        print(f"  CSV salvo: {csv_path}")
        return csv_path

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Leads PGFN SP"

    # Cabecalho
    headers = list(dados[0].keys()) if dados else []
    header_fill = PatternFill("solid", fgColor="1F3864")
    header_font = Font(bold=True, color="FFFFFF", size=10)
    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border

    ws.row_dimensions[1].height = 35

    # Linhas de dados
    alt_fill   = PatternFill("solid", fgColor="EBF1DE")
    money_fmt  = '#,##0.00'
    int_fmt    = '#,##0'

    for row_idx, row_data in enumerate(dados, 2):
        fill = alt_fill if row_idx % 2 == 0 else None
        for col_idx, key in enumerate(headers, 1):
            val = row_data.get(key, "")
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.alignment = Alignment(vertical="top", wrap_text=(key == "Socios"))
            cell.border = border
            if fill:
                cell.fill = fill

            # Formatacao numerica
            if key in ("Divida PGFN (R$)", "Capital Social"):
                cell.number_format = money_fmt
            elif key == "Inscricoes PGFN":
                cell.number_format = int_fmt

    # Ajusta largura das colunas
    col_widths = {
        "CNPJ": 16, "Razao Social": 45, "Nome Fantasia": 25, "CNAE": 8,
        "Setor": 40, "Porte": 12, "Capital Social": 16, "Divida PGFN (R$)": 18,
        "Inscricoes PGFN": 14, "Email": 32, "Telefone": 16, "Site": 30,
        "Cidade": 18, "UF": 4, "Logradouro": 30, "CEP": 10,
        "Abertura": 12, "Socios": 60, "Fonte Contato": 16,
    }
    for col_idx, header in enumerate(headers, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = col_widths.get(header, 15)

    # Freeze header
    ws.freeze_panes = "A2"

    # Aba de estatisticas
    ws2 = wb.create_sheet("Resumo")
    total = len(dados)
    com_email = sum(1 for d in dados if d.get("Email"))
    com_tel   = sum(1 for d in dados if d.get("Telefone"))
    com_site  = sum(1 for d in dados if d.get("Site"))
    me_count  = sum(1 for d in dados if d.get("Porte") == "ME")
    epp_count = sum(1 for d in dados if d.get("Porte") == "EPP")
    mg_count  = sum(1 for d in dados if d.get("Porte") == "Medio/Grande")
    div_media = sum(d.get("Divida PGFN (R$)", 0) for d in dados) / max(total, 1)

    ws2_data = [
        ("Metrica", "Valor"),
        ("Total de empresas", total),
        ("Com email", com_email),
        ("Com telefone", com_tel),
        ("Com site", com_site),
        ("ME", me_count),
        ("EPP", epp_count),
        ("Medio/Grande", mg_count),
        ("Divida media (R$)", div_media),
        ("Data geracao", datetime.now().strftime("%d/%m/%Y %H:%M")),
        ("Fonte", "PGFN Dez/2025 + Receita Federal + OpenCNPJ + Web"),
    ]

    for row in ws2_data:
        ws2.append(row)

    ws2["A1"].font = Font(bold=True)
    ws2["B1"].font = Font(bold=True)
    ws2.column_dimensions["A"].width = 25
    ws2.column_dimensions["B"].width = 40

    wb.save(caminho)
    return caminho


# -----------------------------------------------------------------------------
# MAIN
# -----------------------------------------------------------------------------

async def main():
    t_inicio = time.time()
    print(f"\n{'='*60}")
    print("HERMES -- TESTE COMPLETO PROSPECCAO PGFN SP")
    print(f"{'='*60}")
    print(f"Banco: {CNPJ_DB}")
    print(f"Saida: {SAIDA_XLS}")
    print()

    # 1. Selecao
    con_ro = duckdb.connect(CNPJ_DB, read_only=True)
    empresas = selecionar_100_empresas(con_ro)
    cnpjs = [str(e["cnpj"]) for e in empresas]

    # 2. Socios do DB
    print("\n[2/5] Buscando socios no banco...")
    socios_db = buscar_socios_db(con_ro, cnpjs)
    total_com_socios = sum(1 for v in socios_db.values() if v)
    print(f"  Empresas com socios no DB: {total_com_socios}/{len(cnpjs)}")

    # 3. OpenCNPJ
    print("\n[3/5] Enriquecimento OpenCNPJ (3 APIs em fallback)...")
    sem = asyncio.Semaphore(CONCORRENCIA)
    t = time.time()
    oc_results = await enriquecer_opencnpj_lote(empresas, sem)
    oc_ok = sum(1 for v in oc_results.values() if v.get("email_receita") or v.get("telefone"))
    print(f"  OK em {time.time()-t:.1f}s | {oc_ok}/{len(cnpjs)} com dados uteis")

    # 4. Web scraping
    print("\n[4/5] Busca web (DuckDuckGo -> site/contatos)...")
    t = time.time()
    try:
        web_results = await scraping_lote(empresas)
        web_ok = sum(1 for v in web_results.values() if v.get("site"))
        print(f"  OK em {time.time()-t:.1f}s | {web_ok}/{len(cnpjs)} com site encontrado")
    except Exception as exc:
        print(f"  Web scraping falhou: {exc} -- continuando sem dados web")
        web_results = {}

    con_ro.close()

    # 5. Mescla
    print("\n[5/5] Mesclando dados e exportando...")
    mesclados = []
    for emp in empresas:
        cnpj = str(emp["cnpj"])
        mesclado = mesclar(
            emp=emp,
            opencnpj=oc_results.get(cnpj, {}),
            web=web_results.get(cnpj, {}),
            socios=socios_db.get(cnpj, []),
        )
        mesclados.append(mesclado)

    # Persistencia
    try:
        con_rw = duckdb.connect(CNPJ_DB, read_only=False)
        salvos = persistir(con_rw, mesclados)
        total_enriquecidas = con_rw.execute("SELECT COUNT(*) FROM empresas_enriquecidas").fetchone()[0]
        con_rw.commit()
        con_rw.close()
        print(f"  Persistidos em empresas_enriquecidas: {salvos}")
        print(f"  Total na tabela apos persistencia:   {total_enriquecidas:,}")
    except Exception as exc:
        print(f"  Aviso persistencia: {exc}")

    # Export
    caminho = exportar_excel(mesclados, SAIDA_XLS)
    print(f"\n  Excel gerado: {caminho}")

    # Relatorio final
    elapsed = time.time() - t_inicio
    com_email = sum(1 for d in mesclados if d["Email"])
    com_tel   = sum(1 for d in mesclados if d["Telefone"])
    com_site  = sum(1 for d in mesclados if d["Site"])
    com_socios = sum(1 for d in mesclados if d["Socios"])
    div_media  = sum(d["Divida PGFN (R$)"] for d in mesclados) / len(mesclados)

    print(f"\n{'='*60}")
    print("RESULTADO FINAL")
    print(f"{'='*60}")
    print(f"  Total de empresas:   {len(mesclados)}")
    print(f"  Com email:           {com_email} ({com_email/len(mesclados)*100:.0f}%)")
    print(f"  Com telefone:        {com_tel} ({com_tel/len(mesclados)*100:.0f}%)")
    print(f"  Com site:            {com_site} ({com_site/len(mesclados)*100:.0f}%)")
    print(f"  Com socios:          {com_socios} ({com_socios/len(mesclados)*100:.0f}%)")
    print(f"  Divida media:        R$ {div_media:,.2f}")
    print(f"  Tempo total:         {elapsed:.1f}s")
    print()
    print("PREVIEW (10 primeiras):")
    for d in mesclados[:10]:
        print(f"  {d['CNPJ']} | {d['Razao Social'][:35]:35} | R${d['Divida PGFN (R$)']:>12,.0f} | {d['Email'] or '--':30} | {d['Socios'][:40]}")
    print(f"\nArquivo: {caminho}")


if __name__ == "__main__":
    asyncio.run(main())
